"""Build Lisbon_Marathon_Finish_Plan_v7.xlsx from v6.

Loads v6 and rewrites only what the recalibration changes, so history,
formatting and every untouched sheet survive intact.
"""
import re
import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[0]
SRC = Path("/home/skarambhe/Projects/ai-running-coach/Lisbon_Marathon_Finish_Plan_v6.xlsx")
DST = Path("/home/skarambhe/Projects/ai-running-coach/Lisbon_Marathon_Finish_Plan_v7.xlsx")

shutil.copy(SRC, DST)
wb = openpyxl.load_workbook(DST)

# --- 1. Training Plan: title + revision note ---
tp = wb["Training Plan"]
tp["A1"] = "LISBON MARATHON 2026 — FINISH PLAN (v7)"
tp["A2"] = ("Race: Saturday October 10, 2026  |  Target: 4:45–5:00 (~6:50/km)  "
            "|  10 weeks remaining")
tp["A3"] = (
    "v7 (Jul 30): PACE RECALIBRATION. v6 slowed easy pace to 7:00–7:30 on the "
    "reasoning that easy runs were running at HR 150–155 with Z2 topping out near 149. "
    "That ceiling was wrong: it assumed max HR 190–191 (the 220−age estimate), but Garmin "
    "has max HR 199 configured with RHR 44, putting Zone 2 at 137–152. Runs judged too "
    "fast under v6 were mostly correctly easy — Jul 21/25/28/30 logged 87–100% of time in "
    "Z1+Z2 at 6:21–6:48/km. Easy pace is now expressed as an HR cap with expected pace "
    "banded by distance, because Z2 pace decays with duration (11 km at 7:06/km still hit "
    "HR 157). Goal, MP, race plan and all PF rules unchanged. Weeks 1–18 left as history; "
    "week 19 is in progress so its remaining runs carry the new bands."
)

# --- 2. Pace Guide: HR cap primary, pace banded by duration ---
pg = wb["Pace Guide"]
ZONES = [
    ("Easy 5–6 km", "6:20–6:45/km", "Zone 2 (60-70% max HR), HR ≤150",
     "Conversational. Pace is the RESULT of holding the cap, not the target."),
    ("Easy 7–9 km", "6:40–7:05/km", "Zone 2 (60-70% max HR), HR ≤150",
     "Same effort as above; pace drifts slower as the run gets longer. That is correct."),
    ("Long Run 12 km+", "6:55–7:20/km", "Zone 2 (60-70% max HR), HR ≤150",
     "Start at the slow end. Time on feet is the point. Let pace fall as it needs to."),
    ("Marathon Pace (MP)", "6:45/km", "Zone 2-3 (68-75% max HR)",
     "Unchanged pending the 22 km checkpoint on Sep 19 — no long-run evidence yet."),
    ("Strides", "~5:00/km (100m)", "Zone 4-5",
     "Short 20-sec bursts, smooth. Race week only."),
]
pg["A1"] = "PACE REFERENCE — LISBON FINISH PLAN v7 (HR-capped)"
# Header is row 3; v6 had 4 zone rows (4-7). Write 5 and clear any leftover.
for i, (rt, pace, hr, feel) in enumerate(ZONES):
    r = 4 + i
    pg.cell(row=r, column=1, value=rt)
    pg.cell(row=r, column=2, value=pace)
    pg.cell(row=r, column=3, value=hr)
    pg.cell(row=r, column=4, value=feel)
# v6's block after the zones started at row 8 (blank) — one extra zone row means
# the old blank row 8 is now occupied, so shift the free-text blocks down by one.
# Simplest and safest: rewrite every row from the end of the zone table onward.
END_OF_ZONES = 4 + len(ZONES)  # first free row
# NB: ws.cell(row, col, value=None) is a no-op in openpyxl — None is the
# "don't assign" sentinel, not a value. Must assign .value explicitly to clear.
for r in range(END_OF_ZONES, pg.max_row + 2):
    for c in range(1, 5):
        pg.cell(row=r, column=c).value = None

BLOCKS = [
    ("WHY PACE IS NOT THE TARGET", [
        "• Zone 2 is 137–152 bpm: Garmin has max HR 199 and RHR 44 configured (Karvonen/%HRR).",
        "• v6 assumed max 190–191 and put the Z2 ceiling near 149. That was ~3–5 bpm too low.",
        "• Judge a run on TIME IN ZONE, not average HR. A run can average 151 and still spend",
        "   half its time in Z3 — that happened on Jul 02, Jul 05 and Jul 09.",
        "• Hold the cap and accept whatever pace it gives you. Pace is an output, not an input.",
        "• At this fitness MP and easy pace are close together. Normal — do not force a gap.",
    ]),
    ("LONG RUN PACING RULE", [
        "• First 25%: slowest of the band. If it does not feel too slow, you are going too fast.",
        "• Middle 50%: hold HR ≤150, conversational throughout. Let pace drift slower if it must.",
        "• Last 25%: per session prescription (easy, or MP where listed).",
        "• Gel every 40 min from km 8. Walk 30 sec through every water stop.",
        "• Z2 pace decays with distance: 11 km at 7:06/km still hit HR 157. Expect to slow.",
    ]),
    ("PLANTAR FASCIITIS RULES", [
        "• Track AM first-step pain 0-10 daily. That is the signal — NOT how the run felt.",
        "• Pain quiet during runs is normal with PF and is not clearance to add volume.",
        "• 3 consecutive mornings worse after a volume rise → repeat the prior week, do not progress.",
        "• AM pain ≥6, or any change in how you land → stop running, see the physio.",
        "• Keep the loading protocol going all 12 weeks. It is the treatment, not a warm-up.",
        "• The pace recalibration does not change any of this. Pain gates volume, not HR.",
    ]),
]
row = END_OF_ZONES + 1  # leave one blank row after the table
for title, lines in BLOCKS:
    pg.cell(row=row, column=1, value=title)
    row += 1
    for line in lines:
        pg.cell(row=row, column=1, value=line)
        row += 1
    row += 1  # blank separator

# --- 3. Weekly cells, weeks 20-29 only (1-19 stay as history) ---
EASY_SHORT = "6:20–6:45/km"   # 5-6 km
EASY_MID = "6:40–7:05/km"     # 7-9 km
LONG_BAND = "6:55–7:20/km"    # 12 km+

def recalibrate(text: str) -> str | None:
    """Rewrite a v6 day cell onto the v7 bands. Returns None if unchanged."""
    if not text or "km" not in text:
        return None
    m = re.search(r"(\d+(?:\.\d+)?)\s*km", text)
    if not m:
        return None
    dist = float(m.group(1))
    low = text.lower()

    if low.startswith("easy"):
        band = EASY_SHORT if dist <= 6 else EASY_MID
    elif low.startswith("long"):
        # Short "long" runs sit on the easy bands; 12 km+ gets the long band.
        band = LONG_BAND if dist >= 12 else (EASY_SHORT if dist <= 6 else EASY_MID)
    else:
        return None

    # Replace the existing pace band, keep everything else (MP finishes included).
    new = re.sub(r"@\s*\d:\d{2}\s*[–—-]\s*\d:\d{2}\s*/km", f"@ {band} (HR ≤150)", text)
    if new == text:
        # Long-run cells write the band without an "@" ("Long 6 km @ 7:00–7:20/km"
        # covers most, but guard the ones that differ).
        new = re.sub(r"\d:\d{2}\s*[–—-]\s*\d:\d{2}\s*/km", f"{band} (HR ≤150)", text)
    return new if new != text else None

header_row = None
for i, r in enumerate(tp.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
    if r and isinstance(r[0], str) and r[0].strip().lower() == "week":
        header_row = i
        break

changed = []
for row_cells in tp.iter_rows(min_row=header_row + 1, max_row=tp.max_row):
    wv = row_cells[0].value
    wn = int(str(wv).strip()) if (isinstance(wv, (int, float)) or (isinstance(wv, str) and str(wv).strip().isdigit())) else None
    # Week 19 is the current week — its Saturday long run is still ahead, so it
    # gets the new band too. Weeks 1-18 are genuinely past and stay untouched.
    if wn is None or wn < 19:
        continue
    for col in range(3, 10):  # Mon..Sun
        cell = row_cells[col]
        new = recalibrate(str(cell.value or ""))
        if new:
            changed.append((wn, cell.coordinate, cell.value, new))
            cell.value = new

wb.save(DST)

print(f"Wrote {DST}")
print(f"\nRewrote {len(changed)} day cells (weeks 20+):")
for wn, coord, old, new in changed:
    print(f"  Wk{wn} {coord}")
    print(f"    - {old}")
    print(f"    + {new}")
