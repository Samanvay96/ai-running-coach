import re
from collections.abc import Iterable
from dataclasses import dataclass, field
from datetime import date, timedelta
from pathlib import Path

import openpyxl


@dataclass
class PrescribedRun:
    workout_type: str       # "easy", "tempo", "intervals", "mp_tempo", "long", "rest", "race", "shakeout"
    distance_km: float
    target_pace: str        # e.g. "7:00–7:30/km" — a range where the plan writes one
    description: str        # Full cell text
    finish_km: float = 0.0  # Closing segment, e.g. "last 3 km @ MP (6:45)" -> 3.0
    finish_pace: str = ""   # ...and "6:45/km"

    def pace_brief(self) -> str:
        """Pace prescription in one phrase, including any closing segment.

        v6 writes its key long runs as "Long 22 km — last 3 km @ MP (6:45)": a
        Z2 body with a faster finish. A caller reading only target_pace would
        miss the MP segment, which is the point of those sessions.
        """
        parts = []
        if self.target_pace:
            parts.append(self.target_pace)
        if self.finish_pace:
            parts.append(
                f"last {self.finish_km:g} km @ {self.finish_pace}"
                if self.finish_km else f"finish @ {self.finish_pace}"
            )
        return ", ".join(parts) if parts else "no pace given"


@dataclass
class PaceZone:
    run_type: str
    pace: str
    hr_zone: str
    feel: str


@dataclass
class RaceSplit:
    segment: str
    target_pace: str
    cumulative_time: str


@dataclass
class FuelingItem:
    when: str
    what: str
    notes: str


@dataclass
class Benchmark:
    """A progress checkpoint.

    v5 framed these as time trials (Distance / Target Time / Target Pace / When
    to Test); v6 reframes them as durability checkpoints (Checkpoint / Target /
    Why / When). The columns line up positionally, so one shape covers both —
    the field names follow v6, which is the live plan.
    """
    checkpoint: str
    target: str
    why: str
    when: str


@dataclass
class GuidanceBlock:
    """A free-text rules block from a sheet, e.g. "PLANTAR FASCIITIS RULES".

    These sit below the tabular data as a title row followed by bullet lines.
    v6 carries most of its coaching intent in them — why the paces dropped, the
    long-run pacing rule, the PF gating rules, the one race-day rule — so they
    get parsed and fed to the model rather than dropped on the floor.
    """
    title: str
    lines: list[str] = field(default_factory=list)


@dataclass
class TrainingWeek:
    week_number: int
    dates: str
    start_date: date
    end_date: date
    phase: str
    monday: PrescribedRun
    tuesday: PrescribedRun
    wednesday: PrescribedRun
    thursday: PrescribedRun
    friday: PrescribedRun
    saturday: PrescribedRun
    sunday: PrescribedRun
    weekly_km_target: float
    notes: str

    def day(self, weekday: int) -> PrescribedRun:
        """Return the prescribed slot for a Python weekday (0=Mon ... 6=Sun)."""
        return (
            self.monday, self.tuesday, self.wednesday, self.thursday,
            self.friday, self.saturday, self.sunday,
        )[weekday]

    def run_slots(self) -> list[tuple[int, PrescribedRun]]:
        """Every non-rest slot this week, as (weekday index, run)."""
        return [(i, self.day(i)) for i in range(7) if self.day(i).workout_type != "rest"]


@dataclass
class ResolvedRun:
    """A prescribed slot matched to the date a run actually happened on.

    `prescribed_date` is the plan's own day for the slot; `query_date` is the
    day being asked about. They differ when a run moved — the Saturday long run
    done on Sunday.
    """
    run: PrescribedRun
    prescribed_date: date
    query_date: date

    @property
    def shifted(self) -> bool:
        return self.prescribed_date != self.query_date

    def shift_note(self) -> str:
        """One phrase naming the move, or '' when the run is on its own day."""
        if not self.shifted:
            return ""
        direction = "carried over from" if self.prescribed_date < self.query_date else "pulled forward from"
        return f"{direction} {self.prescribed_date.strftime('%A %b %d')}"


class TrainingPlan:
    def __init__(self, xlsx_path: str):
        self._xlsx_path = xlsx_path
        self._mtime: float = 0.0
        self.weeks: list[TrainingWeek] = []
        self.pace_zones: list[PaceZone] = []
        self.benchmarks: list[Benchmark] = []
        self.race_splits: list[RaceSplit] = []
        self.fueling: list[FuelingItem] = []
        self.guidance: list[GuidanceBlock] = []
        # Plan header metadata (the rows above the week table)
        self.title: str = ""
        self.goal_line: str = ""
        self.revision_note: str = ""
        self.target_finish: str = ""
        self.target_pace: str = ""
        # Phase banner rows, as (first week number under the banner, text)
        self.section_markers: list[tuple[int, str]] = []
        self._parse(xlsx_path)

    def _parse(self, path: str):
        wb = openpyxl.load_workbook(path, data_only=True)
        self.weeks = []
        self.pace_zones = []
        self.benchmarks = []
        self.race_splits = []
        self.fueling = []
        self.guidance = []
        self.section_markers = []
        # Pace guide and race day are parsed BEFORE the training sheet: run
        # cells fall back to them for prescriptions that name a run type without
        # repeating its pace (v6 long runs, the Lisbon race cell).
        self._parse_pace_guide(wb["Pace Guide"])
        if "Benchmarks" in wb.sheetnames:
            self._parse_benchmarks(wb["Benchmarks"])
        self._parse_race_day(wb["Race Day Plan"])
        self._parse_training_sheet(wb["Training Plan"])
        self._mtime = Path(path).stat().st_mtime

    def reload_if_changed(self) -> bool:
        """Reload the plan if the xlsx file has been modified. Returns True if reloaded."""
        try:
            current_mtime = Path(self._xlsx_path).stat().st_mtime
        except OSError:
            return False
        if current_mtime > self._mtime:
            self._parse(self._xlsx_path)
            return True
        return False

    @staticmethod
    def _find_header_row(ws, first_cell: str, max_scan: int | None = None) -> int | None:
        """Row number whose column A equals `first_cell` (case-insensitive)."""
        target = first_cell.strip().lower()
        limit = ws.max_row if max_scan is None else min(ws.max_row, max_scan)
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=limit, values_only=True), start=1):
            cell = row[0] if row else None
            if isinstance(cell, str) and cell.strip().lower() == target:
                return i
        return None

    def _parse_training_sheet(self, ws):
        plan_start_fallback = date(2026, 3, 2)
        # Locate the "Week | Dates | Phase | Mon..." header rather than assuming
        # a fixed row — v6 added a label row above it and a future revision could
        # shift it again.
        header_row = self._find_header_row(ws, "Week", max_scan=30) or 5
        self._parse_plan_meta(ws, header_row)

        pending_markers: list[str] = []
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=False):
            week_val = row[0].value  # Column A
            # Week numbers arrive as strings ('1', '2', ...) or ints depending on
            # how the sheet was authored. Accept either; any other non-empty cell
            # is a phase banner ("PHASE 3: BASE REBUILD — ...") or an inline note
            # ("Apr 06 – Apr 26: TIME OFF"), which we keep as week context.
            if isinstance(week_val, (int, float)):
                week_num = int(week_val)
            elif isinstance(week_val, str) and week_val.strip().isdigit():
                week_num = int(week_val.strip())
            else:
                if isinstance(week_val, str) and week_val.strip():
                    pending_markers.append(week_val.strip())
                continue

            for marker in pending_markers:
                self.section_markers.append((week_num, marker))
            pending_markers.clear()

            dates_str = str(row[1].value or "")
            parsed = self._parse_date_range(dates_str, year=plan_start_fallback.year)
            if parsed:
                week_start, week_end = parsed
            else:
                week_start = plan_start_fallback + timedelta(weeks=week_num - 1)
                week_end = week_start + timedelta(days=6)

            # Columns 3-9 are Mon..Sun (all 7 days listed); column 10 is the
            # weekly km target; column 19 is notes. Runs, strength, PF loading
            # and rest can fall on any day, so strict=True on every slot —
            # anything without an explicit km figure parses as rest.
            self.weeks.append(TrainingWeek(
                week_number=week_num,
                dates=dates_str,
                start_date=week_start,
                end_date=week_end,
                phase=str(row[2].value or ""),
                monday=self._parse_run_cell(str(row[3].value or ""), "monday", strict=True),
                tuesday=self._parse_run_cell(str(row[4].value or ""), "tuesday", strict=True),
                wednesday=self._parse_run_cell(str(row[5].value or ""), "wednesday", strict=True),
                thursday=self._parse_run_cell(str(row[6].value or ""), "thursday", strict=True),
                friday=self._parse_run_cell(str(row[7].value or ""), "friday", strict=True),
                saturday=self._parse_run_cell(str(row[8].value or ""), "saturday", strict=True),
                sunday=self._parse_run_cell(str(row[9].value or ""), "sunday", strict=True),
                weekly_km_target=float(row[10].value or 0),
                notes=str(row[19].value or ""),
            ))

    def _parse_plan_meta(self, ws, header_row: int):
        """Read the title / goal / revision rows that sit above the week table."""
        lines = [
            str(row[0]).strip()
            for row in ws.iter_rows(min_row=1, max_row=max(header_row - 1, 1), values_only=True)
            if row and row[0] and str(row[0]).strip()
        ]
        self.title = lines[0] if lines else ""
        self.goal_line = lines[1] if len(lines) > 1 else ""
        self.revision_note = " ".join(lines[2:])

        # "Target: 4:45–5:00 (~6:50/km)" -> finish "4:45–5:00", pace "6:50/km"
        # "Target Pace: 5:40/km"         -> pace only (the v5 shape)
        m = re.search(r"Target:\s*([^|(]+?)\s*(?:\(|\||$)", self.goal_line)
        self.target_finish = m.group(1).strip() if m else ""
        m = re.search(r"Target(?:\s+Pace)?:[^|]*?~?\s*(\d:\d{2})\s*/km", self.goal_line)
        self.target_pace = f"{m.group(1)}/km" if m else ""

    _FINISH_RE = re.compile(
        r"last\s+(\d+(?:\.\d+)?)\s*km\s*(?:@|at)\s*(?:MP\s*\(\s*)?(\d:\d{2})",
        re.IGNORECASE,
    )

    def _parse_run_cell(self, text: str, day: str, strict: bool = False) -> PrescribedRun:
        if not text or text == "None":
            return PrescribedRun("rest", 0, "", "Rest")

        text_lower = text.lower()

        # Race day — the plan flags races with the 🏁 emoji and/or
        # "MARATHON"/"HALF" in the cell (and may omit a km figure on the
        # Lisbon cell). Detect any of those so race day doesn't fall through
        # to rest under strict mode.
        if "🏁" in text or "race day" in text_lower or "marathon" in text_lower or "half" in text_lower:
            dist = self._extract_distance(text)
            is_half = "half" in text_lower
            if dist == 0:
                # Default: full marathon unless the cell says "half"
                dist = 21.1 if is_half else 42.2
            # Goal-race pace comes from the Race Day Plan's opening split so it
            # tracks the sheet instead of a constant baked in here (which is how
            # v5's 5:40/km outlived the sub-4:00 goal). A tune-up race is a
            # different event — don't stamp the marathon's pace on it.
            pace = "" if is_half else self._race_opening_pace()
            return PrescribedRun("race", dist, pace, text)

        # Shakeout
        if "shakeout" in text_lower:
            dist = self._extract_distance(text)
            return PrescribedRun("shakeout", dist, "", text)

        # A closing fast segment ("— last 3 km @ MP (6:45)") is a separate
        # prescription from the run's body pace; pull it out before reading the
        # body pace so the two don't get conflated.
        finish_km, finish_pace = 0.0, ""
        fm = self._FINISH_RE.search(text)
        body_text = text
        if fm:
            finish_km = float(fm.group(1))
            finish_pace = f"{fm.group(2)}/km"
            body_text = text[: fm.start()] + text[fm.end():]

        dist = self._extract_distance(text)
        pace = self._extract_pace(body_text)

        # Strict mode: only treat as a run if a km distance is present.
        # Cross-training/strength/PF-loading/rest cells (e.g. "F45 Weights",
        # "Foot/calf loading (PF protocol ~20 min)") have no km figure and should
        # fall through to rest so they aren't matched against actual runs.
        if strict and dist == 0:
            return PrescribedRun("rest", 0, "", text)

        # Strip a leading day-label prefix like "Mon:" so keyword detection works.
        keyword_text = re.sub(r"^\s*(mon|tue|wed|thu|fri|sat|sun)[a-z]*\s*:\s*", "", text_lower)

        if keyword_text.startswith("mp tempo"):
            wtype = "mp_tempo"
        elif keyword_text.startswith("intervals"):
            wtype = "intervals"
        elif keyword_text.startswith("tempo"):
            wtype = "tempo"
        elif keyword_text.startswith("long"):
            wtype = "long"
        elif keyword_text.startswith("easy"):
            wtype = "easy"
        else:
            wtype = "easy"

        # v6 writes long runs as "Long 22 km — last 3 km @ MP (6:45)" with no
        # body pace in the cell, because the Pace Guide owns it. Fall back to the
        # matching pace zone so the coach still has a target to compare against.
        if not pace:
            pace = self._zone_pace(wtype)

        return PrescribedRun(wtype, dist, pace, text, finish_km=finish_km, finish_pace=finish_pace)

    _ZONE_KEYWORDS = {
        "easy": ("easy", "recovery"),
        "long": ("long",),
        "mp_tempo": ("marathon pace",),
        "tempo": ("tempo", "threshold"),
        "intervals": ("interval",),
    }

    def _zone_pace(self, workout_type: str) -> str:
        """Pace for a workout type, read off the Pace Guide sheet."""
        for keyword in self._ZONE_KEYWORDS.get(workout_type, ()):
            for pz in self.pace_zones:
                if keyword in pz.run_type.lower():
                    return pz.pace
        return ""

    def _race_opening_pace(self) -> str:
        """Opening-split pace from the Race Day Plan, e.g. '7:00/km'.

        v6's whole race strategy hangs on starting slow, so the opening split —
        not marathon pace — is what the race cell should advertise.
        """
        for split in self.race_splits:
            m = re.search(r"(\d:\d{2})\s*/km", split.target_pace)
            if m:
                return f"{m.group(1)}/km"
        return self._zone_pace("mp_tempo")

    _MONTHS = {m: i + 1 for i, m in enumerate(
        ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    )}

    def _parse_date_range(self, text: str, year: int) -> tuple[date, date] | None:
        """Parse strings like 'Mar 02 – Mar 08' or 'Apr 27 - May 03' into (start, end)."""
        m = re.match(
            r"\s*([A-Z][a-z]{2})\s+(\d{1,2})\s*[–—\-]\s*([A-Z][a-z]{2})\s+(\d{1,2})\s*$",
            text,
        )
        if not m:
            return None
        s_mo, s_dy, e_mo, e_dy = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
        if s_mo not in self._MONTHS or e_mo not in self._MONTHS:
            return None
        start = date(year, self._MONTHS[s_mo], s_dy)
        end_year = year + 1 if self._MONTHS[e_mo] < self._MONTHS[s_mo] else year
        end = date(end_year, self._MONTHS[e_mo], e_dy)
        return start, end

    def _extract_distance(self, text: str) -> float:
        m = re.search(r"(\d+(?:\.\d+)?)\s*km", text)
        return float(m.group(1)) if m else 0

    def get_z2_bounds(self, max_hr: int, rhr: int | None = None) -> tuple[int, int] | None:
        """Return absolute (low, high) BPM bounds for Zone 2.

        Uses the Karvonen / %HRR formula: HR = ((max_hr - rhr) * pct) + rhr.
        If rhr is not provided, falls back to %MaxHR. Percentages come from the
        Pace Guide sheet (e.g. '60-70% max HR' is interpreted as 60-70% HRR).
        """
        for pz in self.pace_zones:
            if "zone 2" in pz.hr_zone.lower() or "easy" in pz.run_type.lower() or "recovery" in pz.run_type.lower():
                pct = self._parse_hr_zone_pct(pz.hr_zone)
                if pct:
                    low_pct, high_pct = pct
                    if rhr is not None:
                        reserve = max_hr - rhr
                        return (
                            int(round(reserve * low_pct / 100 + rhr)),
                            int(round(reserve * high_pct / 100 + rhr)),
                        )
                    return int(round(max_hr * low_pct / 100)), int(round(max_hr * high_pct / 100))
        return None

    @staticmethod
    def _parse_hr_zone_pct(zone_str: str) -> tuple[int, int] | None:
        """Extract (low, high) percentages from strings like 'Zone 2 (60-70% max HR)'."""
        m = re.search(r"(\d{2,3})\s*[-–—]\s*(\d{2,3})\s*%", zone_str)
        if m:
            return int(m.group(1)), int(m.group(2))
        return None

    def _extract_pace(self, text: str) -> str:
        # A range first ("7:00–7:30/km"). Matching the single-pace pattern
        # instead silently reports only the slow end of the band, which is how
        # "Easy 5 km @ 7:00–7:30/km" used to come back as just "7:30/km".
        m = re.search(r"~?(\d:\d{2})\s*[–—-]\s*(\d:\d{2})\s*/km", text)
        if m:
            return f"{m.group(1)}–{m.group(2)}/km"
        # Single pace, ~6:30/km style
        m = re.search(r"~?(\d:\d{2})\s*/km", text)
        if m:
            return f"{m.group(1)}/km"
        # @5:25 style (tempo/interval target pace)
        m = re.search(r"@\s*(\d:\d{2})", text)
        if m:
            return f"{m.group(1)}/km"
        return ""

    def _parse_pace_guide(self, ws):
        # Pace zones live in a contiguous block under the "Run Type" header; walk
        # until the first blank row to stay robust to zones being added or
        # removed (v5 had 7, v6 has 4 after hills and speedwork were cut).
        header_row = self._find_header_row(ws, "Run Type") or 3
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
            run_type = row[0] if len(row) > 0 else None
            if not run_type:
                break
            self.pace_zones.append(PaceZone(
                run_type=str(run_type),
                pace=str(row[1] or "") if len(row) > 1 else "",
                hr_zone=str(row[2] or "") if len(row) > 2 else "",
                feel=str(row[3] or "") if len(row) > 3 else "",
            ))
        self._parse_guidance(ws)

    def _parse_benchmarks(self, ws):
        header_row = (
            self._find_header_row(ws, "Checkpoint")   # v6
            or self._find_header_row(ws, "Distance")  # v5
            or 3
        )
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
            checkpoint = row[0] if len(row) > 0 else None
            if not checkpoint:
                break
            self.benchmarks.append(Benchmark(
                checkpoint=str(checkpoint),
                target=str(row[1] or "") if len(row) > 1 else "",
                why=str(row[2] or "") if len(row) > 2 else "",
                when=str(row[3] or "") if len(row) > 3 else "",
            ))

    def _parse_race_day(self, ws):
        # Header-driven rather than fixed row numbers. v6 moved both tables up a
        # row, and the old ranges (splits 3-11, fuelling 15-21) only lined up
        # with v6 by luck — against v5 they read the "Split" header as a split,
        # dropped the 40–42.2 km split, and took two section-title rows as
        # fuelling entries. Anchoring on the header cell fits either layout.
        for seg, pace, cume in self._read_table(ws, "Split", 3):
            self.race_splits.append(RaceSplit(
                segment=seg, target_pace=pace, cumulative_time=cume,
            ))
        for when, what, notes in self._read_table(ws, "When", 3):
            self.fueling.append(FuelingItem(when=when, what=what, notes=notes))
        self._parse_guidance(ws)

    def _read_table(self, ws, header_first_cell: str, n_cols: int) -> list[tuple[str, ...]]:
        """Rows under a header row, stopping at the first blank or single-cell row.

        A row with only column A filled means we've reached the next section
        title (e.g. "FUELLING STRATEGY"), not another data row.
        """
        header_row = self._find_header_row(ws, header_first_cell)
        if header_row is None:
            return []
        out: list[tuple[str, ...]] = []
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
            cells = [row[i] if i < len(row) else None for i in range(n_cols)]
            if not cells[0] or all(c is None for c in cells[1:]):
                break
            out.append(tuple(str(c or "") for c in cells))
        return out

    def _parse_guidance(self, ws):
        """Collect free-text rules blocks (a single-cell title + bullet lines).

        Skips row 1 (the sheet title) and any block with no body lines — that
        pattern is a label for a following table, not guidance.
        """
        current: GuidanceBlock | None = None

        def flush():
            nonlocal current
            if current and current.lines:
                self.guidance.append(current)
            current = None

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            first = row[0] if row else None
            rest_filled = any(c is not None for c in row[1:]) if row and len(row) > 1 else False
            if first is None or not str(first).strip() or rest_filled:
                # Blank row, or a tabular row — either way, end the block.
                flush()
                continue
            text = str(first).strip()
            if current is None:
                current = GuidanceBlock(title=text)
            else:
                current.lines.append(text)
        flush()

    def get_week_for_date(self, d: date) -> TrainingWeek | None:
        for week in self.weeks:
            if week.start_date <= d <= week.end_date:
                return week
        return None

    def get_prescribed_run(self, d: date) -> PrescribedRun | None:
        """The run prescribed for exactly this date, ignoring any day shift.

        Prospective callers ("what am I meant to run on Thursday?") want this.
        To match a run that already happened — or one being decided on this
        morning — use resolve_run_for_date, which tolerates a moved day.
        """
        week = self.get_week_for_date(d)
        if not week:
            return None
        slot = week.day(d.weekday())
        return slot if slot.workout_type != "rest" else None

    # A run can only stand in for a slot within this many days. Keeps a Sunday
    # run from claiming Thursday's easy slot once Saturday's is spoken for —
    # at that distance it's a different session, not a moved one.
    MAX_SHIFT_DAYS = 2

    def resolve_run_for_date(
        self,
        d: date,
        completed_dates: Iterable[date] | None = None,
    ) -> ResolvedRun | None:
        """Match a date to the plan slot it fulfils, tolerating a shifted day.

        The plan lays out fixed weekdays (Tue/Thu/Sat in v7), but runs move —
        a Saturday long run gets done on Sunday. Resolving strictly by weekday
        drops the prescription for that run and books Saturday as a miss, so:

          1. If `d` has a run of its own, that's the answer.
          2. Otherwise take the nearest non-rest slot in the same plan week,
             within MAX_SHIFT_DAYS, that no other run has already claimed.
             Ties — a rest day flanked by two slots — go to the earlier one;
             carrying a run over is far more common than pulling one forward.

        `completed_dates` is every date already holding a recorded activity;
        those slots are spoken for and can't be matched twice. Callers without
        DB access may omit it, at the cost of possibly matching a slot that was
        in fact already run.

        Returns None outside the plan window, or when no slot is free.
        """
        week = self.get_week_for_date(d)
        if not week:
            return None

        exact = week.day(d.weekday())
        if exact.workout_type != "rest":
            return ResolvedRun(run=exact, prescribed_date=d, query_date=d)

        taken = set(completed_dates or ())
        # Anchor on d's own Monday rather than week.start_date: day() is indexed
        # by Python weekday, so this is the mapping that's guaranteed to agree
        # with the exact-match branch above.
        monday = d - timedelta(days=d.weekday())
        candidates: list[tuple[int, bool, date, PrescribedRun]] = []
        for i, run in week.run_slots():
            slot_date = monday + timedelta(days=i)
            if slot_date == d or slot_date in taken:
                continue
            if not (week.start_date <= slot_date <= week.end_date):
                continue
            gap = abs((slot_date - d).days)
            if gap > self.MAX_SHIFT_DAYS:
                continue
            # Sort key: nearest first, then past before future.
            candidates.append((gap, slot_date > d, slot_date, run))

        if not candidates:
            return None
        _, _, slot_date, run = min(candidates, key=lambda c: (c[0], c[1]))
        return ResolvedRun(run=run, prescribed_date=slot_date, query_date=d)

    def get_section_marker(self, week_number: int) -> str:
        """The phase banner covering a week, e.g. 'PHASE 3: BASE REBUILD — ...'."""
        current = ""
        for start, text in self.section_markers:
            if start <= week_number:
                current = text
        return current

    def get_goal_summary(self) -> str:
        """One line naming the plan's goal, straight from the xlsx.

        Read off the sheet rather than hardcoded, so a revision that changes the
        target (as v6 did, retiring sub-4:00) can't leave a stale goal embedded
        in the coaching prompts.
        """
        if self.target_finish and self.target_pace:
            return f"target {self.target_finish} (~{self.target_pace})"
        if self.target_pace:
            return f"target pace {self.target_pace}"
        return self.goal_line or self.title

    def get_benchmarks_text(self) -> str:
        """Progress checkpoints, formatted for a prompt."""
        return "\n".join(
            f"  {b.checkpoint}: {b.target}"
            + (f" — {b.why}" if b.why else "")
            + (f" [{b.when}]" if b.when else "")
            for b in self.benchmarks
        )

    def get_guidance_text(self) -> str:
        """Every free-text rules block, formatted for a prompt."""
        out: list[str] = []
        for block in self.guidance:
            out.append(f"{block.title}")
            out.extend(f"  {line}" for line in block.lines)
            out.append("")
        return "\n".join(out).rstrip()

    _DAY_LABELS = ("Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun")

    def get_plan_summary(self) -> str:
        total = len(self.weeks)
        lines = [f"{self.title or 'Training plan'} — {total} weeks, {self.get_goal_summary()}", ""]
        lines.append("PACE ZONES:")
        for pz in self.pace_zones:
            lines.append(f"  {pz.run_type}: {pz.pace} | {pz.hr_zone} | {pz.feel}")
        lines.append("")
        lines.append("PROGRESS CHECKS:")
        lines.append(self.get_benchmarks_text())
        lines.append("")
        lines.append("WEEKS:")
        for w in self.weeks:
            run_days = [
                f"{label}={w.day(i).description}"
                for i, label in enumerate(self._DAY_LABELS)
                if w.day(i).workout_type != "rest"
            ]
            lines.append(
                f"  Wk {w.week_number} ({w.phase}): "
                + " | ".join(run_days)
                + f" | Target={w.weekly_km_target}km"
            )
        return "\n".join(lines)

    def get_week_summary(self, week: TrainingWeek) -> str:
        lines = [f"Week {week.week_number} ({week.phase}) — {week.dates}"]
        marker = self.get_section_marker(week.week_number)
        if marker:
            lines.append(marker)
        for i, label in enumerate(self._DAY_LABELS):
            lines.append(f"{label}: {week.day(i).description}")
        lines.append(f"Target: {week.weekly_km_target} km")
        if week.notes:
            lines.append(f"Notes: {week.notes}")
        return "\n".join(lines)
