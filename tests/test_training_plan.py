"""Tests for the xlsx plan parser, written against the v6 rewrite.

The v6 plan (Jul 20) retired the sub-4:00 goal, slowed easy pace to a 7:00–7:30
band, cut hills and speedwork, and moved its long runs to a "Z2 body + MP
finish" shape. Three parser behaviours had to change for that, and each has a
regression test here:

  1. `_extract_pace` returned only the slow end of a range ("7:00–7:30/km" ->
     "7:30/km"), so the coach compared actual pace against one edge of the band.
  2. "Long 22 km — last 3 km @ MP (6:45)" parsed to an empty pace, hiding the
     MP segment that is the point of those sessions.
  3. Race pace was the hardcoded string "5:40/km" — v5's marathon pace, which
     outlived the goal it belonged to.

Plus the sheet-shape changes: header-driven table location (v6 shifted the Race
Day tables up a row), the Benchmarks sheet's new Checkpoint/Target/Why/When
columns, and the free-text guidance blocks that used to be dropped.

Run with: `.venv/bin/python -m pytest tests/test_training_plan.py`
"""

from datetime import date
from pathlib import Path

import openpyxl
import pytest

from src.training_plan import GuidanceBlock, PrescribedRun, TrainingPlan

# The real workbooks hold personal training data and are gitignored, so they
# exist on the runner's machine but not in a fresh clone. Tests that assert what
# the actual plan says skip when the file is absent; the parser-contract tests
# below build their own synthetic workbooks and always run.
V7 = "Lisbon_Marathon_Finish_Plan_v7.xlsx"
V6 = "Lisbon_Marathon_Finish_Plan_v6.xlsx"
V5 = "Lisbon_Marathon_Sub4_Plan_v5.xlsx"


def _load_or_skip(path: str) -> TrainingPlan:
    if not Path(path).exists():
        pytest.skip(f"{path} not present (gitignored — personal training data)")
    return TrainingPlan(path)


@pytest.fixture(scope="module")
def v6() -> TrainingPlan:
    return _load_or_skip(V6)


@pytest.fixture(scope="module")
def v5() -> TrainingPlan:
    """The previous plan, kept parseable so an older file can still be loaded."""
    return _load_or_skip(V5)


def _week(plan: TrainingPlan, number: int):
    week = next((w for w in plan.weeks if w.week_number == number), None)
    assert week is not None, f"week {number} missing from plan"
    return week


# --- 1. Pace ranges survive parsing ---


def test_easy_pace_keeps_the_whole_range(v6):
    """"Easy 5 km @ 7:00–7:30/km" must not collapse to "7:30/km"."""
    tuesday = _week(v6, 19).tuesday
    assert tuesday.workout_type == "easy"
    assert tuesday.target_pace == "7:00–7:30/km"


def test_single_pace_still_parses_without_a_range(v5):
    """The v5 shape ("Easy 4 km ~6:20/km") has one pace and should stay one."""
    assert _week(v5, 2).tuesday.target_pace == "6:20/km"


@pytest.mark.parametrize(
    "text,expected",
    [
        ("Easy 5 km @ 7:00–7:30/km", "7:00–7:30/km"),   # en dash
        ("Easy 5 km @ 7:00-7:30/km", "7:00–7:30/km"),   # hyphen
        ("Easy 5 km @ 7:00 – 7:30/km", "7:00–7:30/km"),  # spaced
        ("Long 9 km ~6:10/km", "6:10/km"),
        ("Tempo 7 km (2 easy + 3 @ 5:15/km + 2 easy)", "5:15/km"),
        ("Tempo 8 km (2 easy + 4 @5:10 + 2 easy)", "5:10/km"),
        ("Foot/calf loading (PF protocol ~20 min)", ""),
        ("REST", ""),
    ],
)
def test_extract_pace_shapes(v6, text, expected):
    assert v6._extract_pace(text) == expected


# --- 2. Closing MP segment on the long runs ---


def test_long_run_mp_finish_is_parsed_separately(v6):
    """Week 26: "Long 22 km — last 3 km @ MP (6:45)"."""
    saturday = _week(v6, 26).saturday
    assert saturday.workout_type == "long"
    assert saturday.distance_km == 22.0
    assert saturday.finish_km == 3.0
    assert saturday.finish_pace == "6:45/km"


def test_long_run_body_pace_falls_back_to_the_pace_guide(v6):
    """The cell gives no body pace, so the Long Run zone supplies it — and the
    MP finish must not be mistaken for the pace of the whole run."""
    saturday = _week(v6, 26).saturday
    assert saturday.target_pace == "7:00–7:20/km"
    assert saturday.pace_brief() == "7:00–7:20/km, last 3 km @ 6:45/km"


def test_peak_week_long_run(v6):
    """Week 27 is the 26 km peak with a 4 km MP finish."""
    saturday = _week(v6, 27).saturday
    assert (saturday.distance_km, saturday.finish_km, saturday.finish_pace) == (26.0, 4.0, "6:45/km")


def test_finish_segment_does_not_steal_the_total_distance(v6):
    """"Long 22 km — last 3 km ..." must report 22 km, not 3 km."""
    assert _week(v6, 26).saturday.distance_km == 22.0


def test_pace_brief_without_any_pace():
    assert PrescribedRun("easy", 5, "", "Easy 5 km").pace_brief() == "no pace given"


def test_pace_brief_with_only_a_finish_segment():
    run = PrescribedRun("long", 20, "", "Long 20 km", finish_km=0, finish_pace="6:45/km")
    assert run.pace_brief() == "finish @ 6:45/km"


# --- 3. Race pace comes from the plan, not a constant ---


def test_marathon_pace_comes_from_the_opening_split(v6):
    """v6 opens at 7:00/km. The old code stamped every race cell "5:40/km"."""
    race = _week(v6, 29).saturday
    assert race.workout_type == "race"
    assert race.distance_km == 42.2
    assert race.target_pace == "7:00/km"


def test_v5_race_pace_also_tracks_its_own_sheet(v5):
    """Same mechanism on the older plan: v5's opening split is 5:45/km."""
    assert _week(v5, 29).saturday.target_pace == "5:45/km"


def test_tuneup_half_does_not_inherit_marathon_pace(v6):
    """Week 17's Battersea half is a different event — no marathon pace on it."""
    half = _week(v6, 17).saturday
    assert half.workout_type == "race"
    assert half.distance_km == 21.1
    assert half.target_pace == ""


def test_race_opening_pace_falls_back_to_mp_zone_when_no_splits(v6):
    plan = _load_or_skip(V6)
    plan.race_splits = []
    assert plan._race_opening_pace() == "6:45/km"  # the MP zone


def test_race_opening_pace_empty_when_nothing_to_fall_back_on(v6):
    plan = _load_or_skip(V6)
    plan.race_splits = []
    plan.pace_zones = []
    assert plan._race_opening_pace() == ""


# --- Plan metadata / goal ---


def test_v6_goal_is_read_off_the_sheet(v6):
    assert v6.target_finish == "4:45–5:00"
    assert v6.target_pace == "6:50/km"
    assert v6.get_goal_summary() == "target 4:45–5:00 (~6:50/km)"


def test_goal_summary_never_mentions_the_retired_target(v6):
    """The whole point of reading the goal from the xlsx."""
    summary = v6.get_goal_summary() + v6.title
    assert "3:57" not in summary
    assert "SUB 4" not in summary.upper()


def test_v5_goal_line_has_a_pace_but_no_finish_time(v5):
    """v5 wrote "Target Pace: 5:40/km" — pace only, no finish time to find."""
    assert v5.target_pace == "5:40/km"
    assert v5.target_finish == ""
    assert v5.get_goal_summary() == "target pace 5:40/km"


def test_revision_note_is_captured(v6):
    assert v6.revision_note.startswith("v6 (Jul 20): REBUILT")
    assert "Hills and all speedwork removed" in v6.revision_note


def test_title_is_captured(v6):
    assert v6.title == "LISBON MARATHON 2026 — FINISH PLAN (v6)"


# --- Phase banners ---


def test_section_markers_map_to_the_weeks_below_them(v6):
    assert v6.get_section_marker(19).startswith("PHASE 3: BASE REBUILD")
    assert v6.get_section_marker(27).startswith("PHASE 4: TAPER & RACE")
    assert v6.get_section_marker(1).startswith("PHASE 1: ADAPTATION")


def test_phase_banner_prefers_the_phase_over_an_inline_note(v6):
    """Week 6 has both a TIME OFF note and the PHASE 1B banner above it."""
    assert v6.get_section_marker(6).startswith("PHASE 1B: RESTART")


def test_section_marker_for_a_week_before_any_banner(v6):
    assert v6.get_section_marker(0) == ""


# --- Benchmarks sheet ---


def test_v6_benchmarks_use_the_new_columns(v6):
    first = v6.benchmarks[0]
    assert first.checkpoint == "AM foot pain trend"
    assert first.target == "Trending down, <3/10"
    assert first.why == "The gating signal for the whole block"
    assert first.when == "Weekly, every week"
    assert len(v6.benchmarks) == 5


def test_v5_benchmarks_map_positionally(v5):
    """v5's Distance/Target Time/Target Pace/When to Test line up in order."""
    first = v5.benchmarks[0]
    assert (first.checkpoint, first.target, first.why) == ("5 km", "< 26:00", "5:12/km")
    assert len(v5.benchmarks) == 3


def test_benchmarks_text_does_not_pass_why_off_as_a_pace(v6):
    """The old formatter rendered "(why) by when" — reading the rationale as a
    target pace. The rendered text should keep them distinguishable."""
    text = v6.get_benchmarks_text()
    assert "AM foot pain trend: Trending down, <3/10 — The gating signal" in text


# --- Race Day sheet: header-driven tables ---


def test_all_race_splits_are_read(v6):
    assert len(v6.race_splits) == 9
    assert v6.race_splits[0].segment == "0–5 km"
    assert v6.race_splits[-1].segment == "40–42.2 km"
    assert v6.race_splits[-1].cumulative_time == "4:52:15"


def test_v5_split_table_is_neither_truncated_nor_headed_by_its_header(v5):
    """Against v5 the old fixed rows 3-11 read the "Split" header row as a split
    and cut the 40–42.2 km row off the end."""
    segments = [s.segment for s in v5.race_splits]
    assert "Split" not in segments
    assert segments[0] == "0–5 km"
    assert segments[-1] == "40–42.2 km"
    assert len(segments) == 9


def test_v5_fueling_is_not_truncated(v5):
    """The old fixed rows 15-21 spent two slots on title rows, losing the last
    two real entries."""
    whens = [f.when for f in v5.fueling]
    assert "Km 32" in whens
    assert whens[-1] == "Every aid station"


def test_fueling_items_exclude_section_titles(v6):
    assert len(v6.fueling) == 7
    whens = [f.when for f in v6.fueling]
    assert whens[0] == "Morning (3 hrs pre)"
    assert whens[-1] == "Km 32 onward"
    assert "FUELLING STRATEGY" not in whens
    assert "When" not in whens


def test_v5_fueling_no_longer_swallows_the_header_rows(v5):
    """The old fixed range started on v5's "FUELLING STRATEGY" title row."""
    whens = [f.when for f in v5.fueling]
    assert "FUELLING STRATEGY" not in whens
    assert "When" not in whens
    assert whens[0] == "Morning (3 hrs pre)"
    assert whens[-1] == "Every aid station"


def test_splits_table_stops_before_the_next_section(v6):
    """No split should have picked up a fuelling row."""
    assert all("Gel" not in s.target_pace for s in v6.race_splits)


# --- Guidance blocks ---


def test_pf_rules_are_parsed(v6):
    block = next((b for b in v6.guidance if "PLANTAR FASCIITIS" in b.title), None)
    assert block is not None
    assert len(block.lines) == 5
    assert any("3 consecutive mornings worse" in line for line in block.lines)
    assert any("AM pain ≥6" in line for line in block.lines)


def test_all_v6_guidance_blocks_are_found(v6):
    titles = [b.title for b in v6.guidance]
    assert titles == [
        "WHY THE PACES DROPPED",
        "LONG RUN PACING RULE",
        "PLANTAR FASCIITIS RULES",
        "THE ONE RULE",
    ]


def test_guidance_skips_labels_that_head_a_table(v6):
    """"FUELLING STRATEGY" is a single-cell row too, but it labels a table."""
    assert all(b.title != "FUELLING STRATEGY" for b in v6.guidance)


def test_guidance_skips_sheet_titles(v6):
    """Row 1 of each sheet is its title, not the head of a rules block."""
    assert all("PACE REFERENCE" not in b.title for b in v6.guidance)
    assert all("RACE DAY PLAN" not in b.title for b in v6.guidance)


def test_guidance_text_includes_titles_and_bullets(v6):
    text = v6.get_guidance_text()
    assert "PLANTAR FASCIITIS RULES" in text
    assert "Hold 7:00. Bank nothing." in text


def test_v5_guidance_still_parses(v5):
    titles = [b.title for b in v5.guidance]
    assert titles == ["LONG RUN PACING RULE"]


def test_guidance_text_is_empty_when_there_are_no_blocks(v6):
    plan = _load_or_skip(V6)
    plan.guidance = []
    assert plan.get_guidance_text() == ""


# --- Pace zones ---


def test_v6_pace_zones_contain_no_speedwork(v6):
    """Hills, tempo and intervals were cut for the plantar fasciitis. If the
    parser invents them the coach will prescribe them."""
    names = " ".join(z.run_type.lower() for z in v6.pace_zones)
    assert "hill" not in names
    assert "tempo" not in names
    assert "interval" not in names
    assert len(v6.pace_zones) == 4


def test_zone_pace_lookup(v6):
    assert v6._zone_pace("easy") == "7:00–7:30/km"
    assert v6._zone_pace("long") == "7:00–7:20/km"
    assert v6._zone_pace("mp_tempo") == "6:45/km"


def test_zone_pace_for_a_type_the_plan_no_longer_has(v6):
    """v6 has no tempo zone — the lookup must come back empty, not guess."""
    assert v6._zone_pace("tempo") == ""
    assert v6._zone_pace("intervals") == ""
    assert v6._zone_pace("nonsense") == ""


def test_z2_bounds_from_the_pace_guide(v6):
    """60-70% HRR off max 191 / RHR 42 — the band the plan cites as "HR <145"."""
    assert v6.get_z2_bounds(191, 42) == (131, 146)


def test_z2_bounds_without_rhr_uses_pct_of_max(v6):
    assert v6.get_z2_bounds(190) == (114, 133)


def test_z2_bounds_none_when_no_zone_has_percentages(v6):
    plan = _load_or_skip(V6)
    plan.pace_zones = []
    assert plan.get_z2_bounds(190, 42) is None


# --- Non-run cells ---


def test_pf_loading_days_are_rest_but_keep_their_text(v6):
    """Foot/calf loading has no km figure, so it must not match against a run —
    but the prescription text still has to reach /today."""
    monday = _week(v6, 19).monday
    assert monday.workout_type == "rest"
    assert monday.distance_km == 0
    assert "Foot/calf loading" in monday.description


def test_get_prescribed_run_returns_none_on_a_loading_day(v6):
    """Week 19 = Jul 27–Aug 02; the Monday is PF loading, not a run."""
    assert v6.get_prescribed_run(date(2026, 7, 27)) is None
    assert v6.get_prescribed_run(date(2026, 7, 28)).workout_type == "easy"


def test_strides_day_still_reports_its_running_distance(v6):
    """Race week: "Easy 4 km + 4×100m strides" is a 4 km run."""
    tuesday = _week(v6, 29).tuesday
    assert tuesday.workout_type == "easy"
    assert tuesday.distance_km == 4.0


def test_out_of_plan_dates(v6):
    assert v6.get_week_for_date(date(2025, 1, 1)) is None
    assert v6.get_prescribed_run(date(2025, 1, 1)) is None


def test_week_dates_are_parsed_not_assumed(v6):
    """Week 19 is Jul 27 – Aug 02, and spans a month boundary."""
    week = _week(v6, 19)
    assert (week.start_date, week.end_date) == (date(2026, 7, 27), date(2026, 8, 2))


def test_weekly_targets_follow_the_v6_rebuild(v6):
    """16 -> 39 km peak, per the revision note."""
    assert _week(v6, 18).weekly_km_target == 16
    assert _week(v6, 27).weekly_km_target == 39
    assert _week(v6, 29).weekly_km_target == 7


def test_week_summary_includes_the_phase_banner(v6):
    summary = v6.get_week_summary(_week(v6, 19))
    assert "PHASE 3: BASE REBUILD" in summary
    assert "Foot/calf loading" in summary
    assert "Target: 19.0 km" in summary


# --- Header-driven location: tolerate a shifted sheet ---


def _write_workbook(path, *, week_header_row: int, benchmarks: bool = True):
    """Build a minimal but structurally faithful plan workbook."""
    wb = openpyxl.Workbook()

    tp = wb.active
    tp.title = "Training Plan"
    tp.cell(row=1, column=1, value="TEST PLAN (v9)")
    tp.cell(row=2, column=1, value="Race: Oct 10, 2026  |  Target: 4:30 (~6:24/km)  |  2 weeks")
    tp.cell(row=3, column=1, value="v9 (Aug 01): test fixture.")
    headers = ["Week", "Dates", "Phase", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun",
               "Weekly km"] + [""] * 8 + ["Notes"]
    for col, name in enumerate(headers, start=1):
        tp.cell(row=week_header_row, column=col, value=name or None)
    r = week_header_row + 1
    tp.cell(row=r, column=1, value="PHASE 1: TESTING (Weeks 1–1)")
    tp.cell(row=r + 1, column=1, value=1)
    tp.cell(row=r + 1, column=2, value="Mar 02 – Mar 08")
    tp.cell(row=r + 1, column=3, value="Test")
    tp.cell(row=r + 1, column=5, value="Easy 5 km @ 7:00–7:30/km")
    tp.cell(row=r + 1, column=9, value="Long 20 km — last 4 km @ MP (6:45)")
    tp.cell(row=r + 1, column=11, value=25)
    tp.cell(row=r + 1, column=20, value="Fixture week.")

    pg = wb.create_sheet("Pace Guide")
    pg.cell(row=1, column=1, value="PACE REFERENCE")
    for col, name in enumerate(["Run Type", "Pace/km", "Heart Rate Zone", "Feel"], start=1):
        pg.cell(row=3, column=col, value=name)
    for i, (rt, pace) in enumerate([("Easy / Recovery", "7:00–7:30/km"),
                                    ("Long Run", "7:00–7:20/km"),
                                    ("Marathon Pace (MP)", "6:45/km")]):
        pg.cell(row=4 + i, column=1, value=rt)
        pg.cell(row=4 + i, column=2, value=pace)
        pg.cell(row=4 + i, column=3, value="Zone 2 (60-70% max HR)")
        pg.cell(row=4 + i, column=4, value="Conversational.")
    pg.cell(row=9, column=1, value="TEST RULES")
    pg.cell(row=10, column=1, value="• Rule one.")
    pg.cell(row=11, column=1, value="• Rule two.")

    rd = wb.create_sheet("Race Day Plan")
    rd.cell(row=1, column=1, value="RACE DAY PLAN")
    for col, name in enumerate(["Split", "Target Pace", "Cumulative Time"], start=1):
        rd.cell(row=2, column=col, value=name)
    rd.cell(row=3, column=1, value="0–5 km")
    rd.cell(row=3, column=2, value="6:30/km (steady)")
    rd.cell(row=3, column=3, value="0:32:30")
    rd.cell(row=5, column=1, value="FUELLING STRATEGY")
    for col, name in enumerate(["When", "What", "Notes"], start=1):
        rd.cell(row=6, column=col, value=name)
    rd.cell(row=7, column=1, value="Km 8")
    rd.cell(row=7, column=2, value="Gel #1")
    rd.cell(row=7, column=3, value="With water.")

    if benchmarks:
        bm = wb.create_sheet("Benchmarks")
        bm.cell(row=1, column=1, value="PROGRESS CHECKS")
        for col, name in enumerate(["Checkpoint", "Target", "Why", "When"], start=1):
            bm.cell(row=3, column=col, value=name)
        bm.cell(row=4, column=1, value="Long run 20 km")
        bm.cell(row=4, column=2, value="Complete")
        bm.cell(row=4, column=3, value="Because.")
        bm.cell(row=4, column=4, value="Wk 1")

    wb.save(path)
    return path


def _write_two_week_workbook(path, *, week_header_row: int = 5):
    """Like `_write_workbook`, but with two consecutive weeks (Tue + Sat runs
    each), so cross-week-boundary shifts — a Saturday long run done the
    following Monday — can be tested."""
    wb = openpyxl.Workbook()

    tp = wb.active
    tp.title = "Training Plan"
    tp.cell(row=1, column=1, value="TEST PLAN (v9)")
    tp.cell(row=2, column=1, value="Race: Oct 10, 2026  |  Target: 4:30 (~6:24/km)  |  2 weeks")
    tp.cell(row=3, column=1, value="v9 (Aug 01): test fixture.")
    headers = ["Week", "Dates", "Phase", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun",
               "Weekly km"] + [""] * 8 + ["Notes"]
    for col, name in enumerate(headers, start=1):
        tp.cell(row=week_header_row, column=col, value=name or None)
    r = week_header_row + 1
    tp.cell(row=r, column=1, value="PHASE 1: TESTING (Weeks 1–2)")
    tp.cell(row=r + 1, column=1, value=1)
    tp.cell(row=r + 1, column=2, value="Mar 02 – Mar 08")
    tp.cell(row=r + 1, column=3, value="Test")
    tp.cell(row=r + 1, column=5, value="Easy 5 km @ 7:00–7:30/km")
    tp.cell(row=r + 1, column=9, value="Long 20 km @ 7:00–7:20/km")
    tp.cell(row=r + 1, column=11, value=25)
    tp.cell(row=r + 2, column=1, value=2)
    tp.cell(row=r + 2, column=2, value="Mar 09 – Mar 15")
    tp.cell(row=r + 2, column=3, value="Test")
    tp.cell(row=r + 2, column=5, value="Easy 5 km @ 7:00–7:30/km")
    tp.cell(row=r + 2, column=9, value="Long 22 km @ 7:00–7:20/km")
    tp.cell(row=r + 2, column=11, value=27)

    pg = wb.create_sheet("Pace Guide")
    pg.cell(row=1, column=1, value="PACE REFERENCE")
    for col, name in enumerate(["Run Type", "Pace/km", "Heart Rate Zone", "Feel"], start=1):
        pg.cell(row=3, column=col, value=name)
    for i, (rt, pace) in enumerate([("Easy / Recovery", "7:00–7:30/km"),
                                    ("Long Run", "7:00–7:20/km"),
                                    ("Marathon Pace (MP)", "6:45/km")]):
        pg.cell(row=4 + i, column=1, value=rt)
        pg.cell(row=4 + i, column=2, value=pace)
        pg.cell(row=4 + i, column=3, value="Zone 2 (60-70% max HR)")
        pg.cell(row=4 + i, column=4, value="Conversational.")

    rd = wb.create_sheet("Race Day Plan")
    rd.cell(row=1, column=1, value="RACE DAY PLAN")
    for col, name in enumerate(["Split", "Target Pace", "Cumulative Time"], start=1):
        rd.cell(row=2, column=col, value=name)
    rd.cell(row=3, column=1, value="0–5 km")
    rd.cell(row=3, column=2, value="6:30/km (steady)")
    rd.cell(row=3, column=3, value="0:32:30")
    rd.cell(row=5, column=1, value="FUELLING STRATEGY")
    for col, name in enumerate(["When", "What", "Notes"], start=1):
        rd.cell(row=6, column=col, value=name)
    rd.cell(row=7, column=1, value="Km 8")
    rd.cell(row=7, column=2, value="Gel #1")
    rd.cell(row=7, column=3, value="With water.")

    wb.save(path)
    return path


@pytest.mark.parametrize("header_row", [5, 4, 8])
def test_week_table_is_found_wherever_its_header_sits(tmp_path, header_row):
    """v6 added a label row above the week header; a fixed min_row=5 would break
    on any further shift."""
    path = _write_workbook(tmp_path / f"plan{header_row}.xlsx", week_header_row=header_row)
    plan = TrainingPlan(str(path))
    assert len(plan.weeks) == 1
    week = plan.weeks[0]
    assert week.week_number == 1
    assert week.tuesday.target_pace == "7:00–7:30/km"
    assert week.saturday.finish_pace == "6:45/km"
    assert week.weekly_km_target == 25
    assert week.notes == "Fixture week."
    assert plan.get_section_marker(1) == "PHASE 1: TESTING (Weeks 1–1)"


def test_synthetic_plan_metadata(tmp_path):
    plan = TrainingPlan(str(_write_workbook(tmp_path / "p.xlsx", week_header_row=5)))
    assert plan.target_finish == "4:30"
    assert plan.target_pace == "6:24/km"
    assert plan.get_goal_summary() == "target 4:30 (~6:24/km)"


def test_synthetic_race_and_fuel_tables_stay_separate(tmp_path):
    """The fuelling table sits two rows below the splits with a title between —
    neither may bleed into the other."""
    plan = TrainingPlan(str(_write_workbook(tmp_path / "p.xlsx", week_header_row=5)))
    assert [s.segment for s in plan.race_splits] == ["0–5 km"]
    assert [f.when for f in plan.fueling] == ["Km 8"]


def test_missing_benchmarks_sheet_is_tolerated(tmp_path):
    path = _write_workbook(tmp_path / "nobm.xlsx", week_header_row=5, benchmarks=False)
    plan = TrainingPlan(str(path))
    assert plan.benchmarks == []
    assert plan.get_benchmarks_text() == ""
    assert len(plan.weeks) == 1


def test_read_table_returns_empty_when_the_header_is_absent(tmp_path):
    plan = TrainingPlan(str(_write_workbook(tmp_path / "p.xlsx", week_header_row=5)))
    wb = openpyxl.load_workbook(str(_write_workbook(tmp_path / "p2.xlsx", week_header_row=5)))
    assert plan._read_table(wb["Race Day Plan"], "NoSuchHeader", 3) == []


def test_find_header_row_is_case_insensitive(tmp_path):
    plan = TrainingPlan(str(_write_workbook(tmp_path / "p.xlsx", week_header_row=5)))
    wb = openpyxl.load_workbook(str(_write_workbook(tmp_path / "p2.xlsx", week_header_row=5)))
    assert plan._find_header_row(wb["Pace Guide"], "run type") == 3
    assert plan._find_header_row(wb["Pace Guide"], "missing") is None


# --- Reload ---


def test_reload_picks_up_an_edit(tmp_path):
    path = _write_workbook(tmp_path / "p.xlsx", week_header_row=5)
    plan = TrainingPlan(str(path))
    assert plan.weeks[0].weekly_km_target == 25

    wb = openpyxl.load_workbook(str(path))
    wb["Training Plan"].cell(row=6 + 1, column=11, value=30)
    # mtime has 1s granularity on some filesystems; force a distinct value.
    wb.save(str(path))
    import os
    stat = os.stat(path)
    os.utime(path, (stat.st_atime + 10, stat.st_mtime + 10))

    assert plan.reload_if_changed() is True
    assert plan.weeks[0].weekly_km_target == 30
    assert plan.reload_if_changed() is False


def test_reload_on_a_vanished_file_returns_false(tmp_path):
    path = _write_workbook(tmp_path / "gone.xlsx", week_header_row=5)
    plan = TrainingPlan(str(path))
    path.unlink()
    assert plan.reload_if_changed() is False
    # The already-parsed plan stays usable rather than blanking out.
    assert len(plan.weeks) == 1


def test_reload_does_not_duplicate_parsed_rows(tmp_path):
    """_parse resets every collection; a reload must not append a second copy."""
    path = _write_workbook(tmp_path / "p.xlsx", week_header_row=5)
    plan = TrainingPlan(str(path))
    before = (len(plan.weeks), len(plan.pace_zones), len(plan.race_splits),
              len(plan.fueling), len(plan.guidance), len(plan.section_markers),
              len(plan.benchmarks))
    plan._parse(str(path))
    after = (len(plan.weeks), len(plan.pace_zones), len(plan.race_splits),
             len(plan.fueling), len(plan.guidance), len(plan.section_markers),
             len(plan.benchmarks))
    assert before == after


# --- Empty / defensive cell handling ---


def test_blank_cell_is_rest(v6):
    assert v6._parse_run_cell("", "monday").workout_type == "rest"
    assert v6._parse_run_cell("None", "monday").workout_type == "rest"


def test_strict_mode_rejects_a_cell_with_no_distance(v6):
    run = v6._parse_run_cell("REST or bodyweight", "monday", strict=True)
    assert run.workout_type == "rest"
    assert run.description == "REST or bodyweight"


def test_non_strict_mode_keeps_a_distanceless_run(v6):
    """Used by callers that already know the cell is a run."""
    assert v6._parse_run_cell("Easy run, no distance given", "tuesday").workout_type == "easy"


def test_guidance_block_defaults_to_no_lines():
    assert GuidanceBlock(title="X").lines == []


# --- v7: the HR-capped pace recalibration ---
#
# v7 replaces v6's single 7:00-7:30 easy band with an HR cap plus pace bands that
# widen with distance, because Zone 2 pace decays with duration (11 km at 7:06/km
# still hit HR 157). The band itself was also wrong: v6 sized Zone 2 off max HR
# 190-191 where Garmin has 199 configured.


@pytest.fixture(scope="module")
def v7() -> TrainingPlan:
    return _load_or_skip(V7)


def test_v7_pace_zones_are_banded_by_duration(v7):
    bands = {z.run_type: z.pace for z in v7.pace_zones}
    assert bands["Easy 5–6 km"] == "6:20–6:45/km"
    assert bands["Easy 7–9 km"] == "6:40–7:05/km"
    assert bands["Long Run 12 km+"] == "6:55–7:20/km"


def test_v7_keeps_mp_and_strides_unchanged(v7):
    """The goal and MP were deliberately left alone pending the Sep 19 checkpoint."""
    bands = {z.run_type: z.pace for z in v7.pace_zones}
    assert bands["Marathon Pace (MP)"] == "6:45/km"
    assert bands["Strides"] == "~5:00/km (100m)"
    assert v7.target_finish == "4:45–5:00"
    assert v7.target_pace == "6:50/km"


def test_v7_every_easy_zone_names_the_hr_cap(v7):
    """Pace is the output; the cap is the actual prescription."""
    for z in v7.pace_zones:
        if z.run_type.lower().startswith(("easy", "long")):
            assert "HR ≤150" in z.hr_zone


def test_v7_z2_fallback_band_matches_garmin(v7):
    """Even the plan-derived fallback should now land on 137-152, because the
    Pace Guide percentages are applied to the correct max HR."""
    assert v7.get_z2_bounds(199, 44) == (137, 152)


def test_v7_short_easy_runs_use_the_faster_band(v7):
    week = _week(v7, 21)
    assert week.tuesday.distance_km == 6.0
    assert week.tuesday.target_pace == "6:20–6:45/km"


def test_v7_mid_distance_easy_runs_use_the_middle_band(v7):
    week = _week(v7, 24)
    assert week.tuesday.distance_km == 7.0
    assert week.tuesday.target_pace == "6:40–7:05/km"


def test_v7_long_runs_use_the_long_band(v7):
    week = _week(v7, 24)
    assert week.saturday.distance_km == 18.0
    assert week.saturday.target_pace == "6:55–7:20/km"


def test_v7_mp_finish_long_runs_survive_the_rewrite(v7):
    """Weeks 26/27 carry no explicit band, so they fall back to the Long Run
    zone — and the MP closing segment must still parse."""
    saturday = _week(v7, 26).saturday
    assert saturday.pace_brief() == "6:55–7:20/km, last 3 km @ 6:45/km"
    peak = _week(v7, 27).saturday
    assert (peak.distance_km, peak.finish_km, peak.finish_pace) == (26.0, 4.0, "6:45/km")


def test_v7_current_week_is_recalibrated(v7):
    """Week 19 is in progress — its remaining Saturday long run must carry the
    new band, not a stale one."""
    week = _week(v7, 19)
    assert week.tuesday.target_pace == "6:20–6:45/km"
    assert week.saturday.target_pace == "6:40–7:05/km"


def test_v7_leaves_completed_weeks_untouched(v7, v6):
    """Weeks 1-18 are history and must read exactly as v6 wrote them."""
    for n in (5, 9, 13, 17, 18):
        old, new = _week(v6, n), _week(v7, n)
        for i in range(7):
            assert old.day(i).description == new.day(i).description, f"week {n} day {i}"


def test_v7_preserves_the_schedule_and_volume(v7, v6):
    """Only paces changed — distances, targets and rest days are v6's."""
    for a, b in zip(v6.weeks, v7.weeks):
        assert a.week_number == b.week_number
        assert a.weekly_km_target == b.weekly_km_target
        for i in range(7):
            assert a.day(i).distance_km == b.day(i).distance_km
            assert a.day(i).workout_type == b.day(i).workout_type


def test_v7_pf_rules_are_untouched(v7):
    """The foot is gated by AM pain, not by heart rate — this recalibration
    must not have loosened anything there."""
    block = next(b for b in v7.guidance if "PLANTAR FASCIITIS" in b.title)
    assert any("3 consecutive mornings worse" in l for l in block.lines)
    assert any("AM pain ≥6" in l for l in block.lines)
    assert any("Pain gates volume, not HR" in l for l in block.lines)


def test_v7_explains_the_max_hr_correction(v7):
    """The rationale block has to state the actual error, since it is what the
    coach cites when the runner asks why the paces moved again."""
    block = next(b for b in v7.guidance if "WHY PACE IS NOT THE TARGET" in b.title)
    text = " ".join(block.lines)
    assert "137–152" in text
    assert "199" in text
    assert "TIME IN ZONE" in text


def test_v7_retires_the_stale_rationale(v7):
    """v6's 'WHY THE PACES DROPPED' was built on the wrong ceiling — it must be
    gone, not sitting alongside the corrected block."""
    assert all("WHY THE PACES DROPPED" not in b.title for b in v7.guidance)


def test_v7_guidance_blocks(v7):
    assert [b.title for b in v7.guidance] == [
        "WHY PACE IS NOT THE TARGET",
        "LONG RUN PACING RULE",
        "PLANTAR FASCIITIS RULES",
        "THE ONE RULE",
    ]


def test_v7_race_day_and_benchmarks_unchanged(v7, v6):
    assert [s.segment for s in v7.race_splits] == [s.segment for s in v6.race_splits]
    assert [s.target_pace for s in v7.race_splits] == [s.target_pace for s in v6.race_splits]
    assert [f.when for f in v7.fueling] == [f.when for f in v6.fueling]
    assert [b.checkpoint for b in v7.benchmarks] == [b.checkpoint for b in v6.benchmarks]


def test_v7_revision_note_records_the_recalibration(v7):
    assert v7.revision_note.startswith("v7 (Jul 30)")
    assert "199" in v7.revision_note
    assert v7.title.endswith("(v7)")


# --- Shift-tolerant slot resolution ---
#
# The plan pins each session to a weekday (Tue/Thu/Sat in v7), but runs move: a
# Saturday long run gets done on Sunday. Matching strictly by weekday threw the
# prescription away for that run and booked Saturday as a miss, so
# resolve_run_for_date matches a run to the nearest free slot in its plan week.
# The synthetic fixture week is Mon 2026-03-02 – Sun 03-08, running Tue + Sat.


@pytest.fixture
def shift_plan(tmp_path) -> TrainingPlan:
    path = _write_workbook(tmp_path / "shift.xlsx", week_header_row=5)
    return TrainingPlan(str(path))


def test_resolve_returns_the_days_own_slot_unshifted(shift_plan):
    r = shift_plan.resolve_run_for_date(date(2026, 3, 7))  # Saturday
    assert r is not None
    assert r.run.workout_type == "long"
    assert r.prescribed_date == date(2026, 3, 7)
    assert r.shifted is False
    assert r.shift_note() == ""


def test_sunday_run_resolves_to_saturdays_long_run(shift_plan):
    """The case this exists for: Saturday's long run done on Sunday."""
    r = shift_plan.resolve_run_for_date(date(2026, 3, 8))
    assert r is not None
    assert r.run.workout_type == "long"
    assert r.run.distance_km == 20.0
    assert r.prescribed_date == date(2026, 3, 7)
    assert r.shifted is True
    assert r.shift_note() == "carried over from Saturday Mar 07"


def test_sunday_run_does_not_claim_a_slot_saturday_already_filled(shift_plan):
    """Ran Saturday AND Sunday → Sunday is a genuine extra, not the long run."""
    r = shift_plan.resolve_run_for_date(
        date(2026, 3, 8), completed_dates={date(2026, 3, 7)}
    )
    assert r is None


def test_friday_run_pulls_saturdays_long_run_forward(shift_plan):
    r = shift_plan.resolve_run_for_date(date(2026, 3, 6))
    assert r is not None
    assert r.run.workout_type == "long"
    assert r.prescribed_date == date(2026, 3, 7)
    assert r.shift_note() == "pulled forward from Saturday Mar 07"


def test_tie_between_flanking_slots_goes_to_the_earlier_one(shift_plan):
    """Wednesday sits one day from Tuesday's slot and (in this fixture) nothing
    else within reach — carrying over beats pulling forward on ties."""
    r = shift_plan.resolve_run_for_date(date(2026, 3, 4))
    assert r is not None
    assert r.prescribed_date == date(2026, 3, 3)
    assert r.run.workout_type == "easy"


def test_resolution_will_not_reach_further_than_max_shift_days(shift_plan):
    """Monday is 4 days from Saturday's slot — too far to be the same session."""
    assert shift_plan.MAX_SHIFT_DAYS == 2
    r = shift_plan.resolve_run_for_date(
        date(2026, 3, 2), completed_dates={date(2026, 3, 3)}
    )
    assert r is None


def test_resolution_stays_inside_the_plan_window(shift_plan):
    """A date the plan doesn't cover has no slot to borrow."""
    assert shift_plan.resolve_run_for_date(date(2026, 3, 1)) is None
    assert shift_plan.resolve_run_for_date(date(2026, 3, 9)) is None


def test_monday_run_resolves_to_prior_weeks_saturday_long_run(tmp_path):
    """A shift can cross the week boundary: Saturday's long run done the
    following Monday is still a 2-day carry-over, not a miss for one plan
    week and an unprescribed extra for the next.

    Week 2's own Tuesday slot (one day out, closer than Saturday's two) is
    marked already-fulfilled, the way it would be if it held its own real
    run — otherwise the nearer slot wins the tie-break, same as any other
    "nearest first" resolution."""
    plan = TrainingPlan(str(_write_two_week_workbook(tmp_path / "two_week.xlsx")))
    r = plan.resolve_run_for_date(date(2026, 3, 9), completed_dates={date(2026, 3, 10)})
    assert r is not None
    assert r.run.workout_type == "long"
    assert r.run.distance_km == 20.0  # week 1's long run, not week 2's
    assert r.prescribed_date == date(2026, 3, 7)  # Saturday, week 1
    assert r.shift_note() == "carried over from Saturday Mar 07"


def test_get_prescribed_run_stays_strict(shift_plan):
    """Prospective callers must keep seeing Sunday as a rest day — otherwise
    "what's coming up" smears every run across its neighbouring days."""
    assert shift_plan.get_prescribed_run(date(2026, 3, 8)) is None
    assert shift_plan.get_prescribed_run(date(2026, 3, 7)).workout_type == "long"
